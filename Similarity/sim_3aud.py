import librosa
import time
import os
import librosa.display
import matplotlib.pyplot as plt
import numpy
!pip install dtw
from dtw import dtw
from numpy.linalg import norm
import statistics as stat
#!pip install mlpy
#import mlpy

from google.colab import drive
#drive.flush_and_unmount()
drive.mount("/content/gdrive", force_remount=True)

#TO CREATE A NOTEBOOK
from openpyxl import Workbook

#PATH OF NOTEBOOK
basepath = "/content/gdrive/My Drive/PAS/"
test_filename = 'sim-record2.xlsx'
wb_path = basepath + "Similarity/" +test_filename

# CREATING IT
wb = Workbook()

ws = wb.active
ws['A1'] = 'SOURCE'
ws['B1'] = 'TARGET'
ws['C1'] = 'DISTANCE'
ws['D1'] = 'DECREMENT'
wb.save(wb_path)

#FOR ONE BASE AND ALL TARGET
start_time = time.time()

basepath = "/content/gdrive/My Drive/PAS/"
result_path = basepath + "results/"

record_filename = 'sim-record1.xlsx'
wb_path = basepath + "Similarity/" + record_filename

import openpyxl
wb = openpyxl.load_workbook(wb_path)
ws=wb.active

#THREE SOURCE FILES
audio1_source_name = "Marina/"
audio_clip1 = "MS-1.wav"
audio1 = basepath + audio1_source_name + audio_clip1

audio2_source_name = "Base/"
audio_clip2 = "BASE-MS-1.WAV"
audio2 = basepath + audio2_source_name + audio_clip2

# audio2_source_name = "Misc/"
# audio_clip2 = "F-O-1.wav"
# audio2 = basepath + audio2_source_name + audio_clip2

audio_clip3 = "MS-1_10000_0.23(17).wav"
audio3 = result_path + audio_clip3

#Loading audio files
y1, sr1 = librosa.load(audio1) 
y2, sr2 = librosa.load(audio2) 
y3, sr3 = librosa.load(audio3)

mfcc1 = librosa.feature.mfcc(y1,sr1)   #Computing MFCC values
mfcc2 = librosa.feature.mfcc(y2, sr2)
mfcc3 = librosa.feature.mfcc(y3, sr3)

#NORMALIZED DISTANCES
distA, cost, acc_cost, path = dtw(mfcc1.T, mfcc2.T, dist=lambda x, y: norm(x - y, ord=1))
print ('Normalized distance between ', audio_clip1, ' and ', audio_clip2, ' is ', distA) # 0 for similar audios
ws.append([audio_clip1, audio_clip2, distA])

distB, cost, acc_cost, path = dtw(mfcc1.T, mfcc3.T, dist=lambda x, y: norm(x - y, ord=1))
print ('Normalized distance between ', audio_clip1, ' and ', audio_clip3, ' is ', distB) # 0 for similar audios
ws.append([audio_clip1, audio_clip3, distB])

# distC, cost, acc_cost, path = dtw(mfcc2.T, mfcc3.T, dist=lambda x, y: norm(x - y, ord=1))
# print ('Normalized distance between ', audio_clip2, ' and ', audio_clip3, ' is ', distC) # 0 for similar audios
# ws.append([audio_clip2, audio_clip3, distC, distA-distC])

wb.save(wb_path)

