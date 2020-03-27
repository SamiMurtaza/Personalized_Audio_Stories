import librosa
import time
import os
import librosa.display
import matplotlib.pyplot as plt
import numpy
!pip install dtw
from dtw import dtw
from numpy.linalg import norm
import statistics as stat
#!pip install mlpy
#import mlpy

from google.colab import drive
#drive.flush_and_unmount()
drive.mount("/content/gdrive", force_remount=True)

#TO CREATE A NOTEBOOK
from openpyxl import Workbook

#PATH OF NOTEBOOK
basepath = "/content/gdrive/My Drive/PAS/"
test_filename = 'sim-record2.xlsx'
wb_path = basepath + "Similarity/" +test_filename

# CREATING IT
wb = Workbook()

ws = wb.active
ws['A1'] = 'SOURCE'
ws['B1'] = 'TARGET'
ws['C1'] = 'DISTANCE'
ws['D1'] = 'DECREMENT'
wb.save(wb_path)

#FOR ONE BASE AND 5 TARGET

basepath = "/content/gdrive/My Drive/PAS/"
result_path = basepath + "results/"

record_filename = 'sim-record3.xlsx'
wb_path = basepath + "Similarity/" + record_filename


import openpyxl
wb = openpyxl.load_workbook(wb_path)
ws=wb.active

#THE ONE BASE AUDIO
audio1_source_name = "Marina/"
audio_clip_s = "MS-1.WAV"
audio_s = basepath + audio1_source_name + audio_clip_s

#FIVE TARGET ONES
target_name = "Maryam/"

audio_clip1 = "SF-2.WAV"
audio1 = basepath + target_name + audio_clip1

audio_clip2 = "SF-26.WAV"
audio2 = basepath + target_name + audio_clip2

audio_clip3 = "SF-14.WAV"
audio3 = basepath + target_name + audio_clip3

audio_clip4 = "SF-15.WAV"
audio4 = basepath + target_name + audio_clip4

audio_clip5 = "SF-16.WAV"
audio5 = basepath + target_name + audio_clip5

target_file_names = [audio_clip1, audio_clip2, audio_clip3, audio_clip4, audio_clip5]

#MIMICKED AUDIO FILE
audio_clip_m = "MS-17_MaryamMaMaryam_10000_0.002/(10000).wav"
audio_m = result_path + audio_clip_m

#Loading audio files
y_s, sr_s = librosa.load(audio_s) 
y1, sr1 = librosa.load(audio1) 
y2, sr2 = librosa.load(audio2) 
y3, sr3 = librosa.load(audio3)
y4, sr4 = librosa.load(audio4)
y5, sr5 = librosa.load(audio5)
y_m, sr_m = librosa.load(audio_m)

#Showing multiple plots using subplot
mfcc_s = librosa.feature.mfcc(y_s, sr_s)

mfcc1 = librosa.feature.mfcc(y1,sr1)   #Computing MFCC values


mfcc2 = librosa.feature.mfcc(y2, sr2)


mfcc3 = librosa.feature.mfcc(y3, sr3)


mfcc4 = librosa.feature.mfcc(y4, sr4)


mfcc5 = librosa.feature.mfcc(y5, sr5)


mfcc_m = librosa.feature.mfcc(y_m, sr_m)


#DISTANCE BETWEEN 1 SOURCE AND 5 TARGET
s_t_dist = []
distA, cost, acc_cost, path = dtw(mfcc_s.T, mfcc1.T, dist=lambda x, y: norm(x - y, ord=1))
s_t_dist.append(distA)

distB, cost, acc_cost, path = dtw(mfcc_s.T, mfcc2.T, dist=lambda x, y: norm(x - y, ord=1))
s_t_dist.append(distB)

distC, cost, acc_cost, path = dtw(mfcc_s.T, mfcc3.T, dist=lambda x, y: norm(x - y, ord=1))
s_t_dist.append(distC)

distD, cost, acc_cost, path = dtw(mfcc_s.T, mfcc4.T, dist=lambda x, y: norm(x - y, ord=1))
s_t_dist.append(distD)

distE, cost, acc_cost, path = dtw(mfcc_s.T, mfcc5.T, dist=lambda x, y: norm(x - y, ord=1))
s_t_dist.append(distE)

s_t_median = stat.median(s_t_dist) 
print(s_t_dist)
print(s_t_median)

print ('Normalized distance between ', audio_clip_s, ' and ', target_name, ' is ', s_t_median) # 0 for similar audios
ws.append([audio_clip_s, target_name, s_t_median, audio_clip1, audio_clip2, audio_clip3, audio_clip4, audio_clip5])

#DISTANCE BETWEEN SOURCE AND MIMICKED
distX, cost, acc_cost, path = dtw(mfcc_s.T, mfcc_m.T, dist=lambda x, y: norm(x - y, ord=1))
print ('Normalized distance between ', audio_clip_s, ' and ', audio_clip_m, ' is ', distX) # 0 for similar audios
ws.append([audio_clip_s, audio_clip_m, distX])

#DISTANCE BETWEEN 5 TARGET AND 1 MIMICKED
t_m_dist = []
dist1, cost, acc_cost, path = dtw(mfcc_m.T, mfcc1.T, dist=lambda x, y: norm(x - y, ord=1))
t_m_dist.append(dist1)

dist2, cost, acc_cost, path = dtw(mfcc_m.T, mfcc2.T, dist=lambda x, y: norm(x - y, ord=1))
t_m_dist.append(dist2)

dist3, cost, acc_cost, path = dtw(mfcc_m.T, mfcc3.T, dist=lambda x, y: norm(x - y, ord=1))
t_m_dist.append(dist3)

dist4, cost, acc_cost, path = dtw(mfcc_m.T, mfcc4.T, dist=lambda x, y: norm(x - y, ord=1))
t_m_dist.append(dist4)

dist5, cost, acc_cost, path = dtw(mfcc_m.T, mfcc5.T, dist=lambda x, y: norm(x - y, ord=1))
t_m_dist.append(dist5)

t_m_median = stat.median(t_m_dist) 
print(t_m_dist)
print(t_m_median)

print ('Normalized distance between ', audio_clip_m, ' and ', target_name, ' is ', t_m_median) # 0 for similar audios
ws.append([audio_clip_m, target_name, t_m_median])

wb.save(wb_path)
