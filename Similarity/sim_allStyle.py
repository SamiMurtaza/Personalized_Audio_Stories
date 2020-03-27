import librosa
import time
import os
import librosa.display
import matplotlib.pyplot as plt
import numpy
!pip install dtw
from dtw import dtw
from numpy.linalg import norm
import statistics as stat
#!pip install mlpy
#import mlpy

from google.colab import drive
#drive.flush_and_unmount()
drive.mount("/content/gdrive", force_remount=True)

#TO CREATE A NOTEBOOK
from openpyxl import Workbook

#PATH OF NOTEBOOK
basepath = "/content/gdrive/My Drive/PAS/"
test_filename = 'sim-record1.xlsx'
wb_path = basepath + "Similarity/" +test_filename

# CREATING IT
wb = Workbook()

ws = wb.active
ws['A1'] = 'SOURCE'
ws['B1'] = 'TARGET'
ws['C1'] = 'DISTANCE'
ws['D1'] = 'DECREMENT'
wb.save(wb_path)

#FOR ONE BASE AND ALL TARGET
start_time = time.time()

basepath = "/content/gdrive/My Drive/PAS/"
result_path = basepath + "results/"

record_filename = 'sim-record1.xlsx'
wb_path = basepath + "Similarity/" + record_filename

import openpyxl
wb = openpyxl.load_workbook(wb_path)
ws=wb.active

#THE ONE BASE AUDIO
audio1_source_name = "Marina/"
audio_clip_s = "MS-17.WAV"
audio_s = basepath + audio1_source_name + audio_clip_s

#MIMICKED AUDIO FILE
audio_clip_m = "MS-19_Marina_10000_0.002/(10000).wav"
audio_m = result_path + audio_clip_m

#Loading audio files
y_s, sr_s = librosa.load(audio_s) 
y_m, sr_m = librosa.load(audio_m)

#Computing MFCC
mfcc_s = librosa.feature.mfcc(y_s, sr_s)
mfcc_m = librosa.feature.mfcc(y_m, sr_m)

#ALL TARGET ONES
tar_name = 'Marina'
tar_mfcc = []
s_t_dist = []
t_m_dist = []

#ALL TARGET AGAINST ONE BASE AND ONE MIMICKED
for i in os.listdir(basepath + tar_name + "/"):
  y, sr = librosa.load(basepath+tar_name+"/"+i)
  mfcc = librosa.feature.mfcc(y,sr)

  distA, cost, acc_cost, path = dtw(mfcc_s.T, mfcc.T, dist=lambda x, y: norm(x - y, ord=1))
  s_t_dist.append(distA)

  dist1, cost, acc_cost, path = dtw(mfcc_m.T, mfcc.T, dist=lambda x, y: norm(x - y, ord=1))
  t_m_dist.append(dist1)

#MEDIAN OF SOURCE TARGET DISTANCE
s_t_median = stat.mean(s_t_dist) 
print(s_t_dist)

#MEDIAN OF TARGET_MIMICKED DISTANCE
t_m_median = stat.mean(t_m_dist) 
print(t_m_dist)

#DISTANCE BETWEEN 1 SOURCE AND ALL TARGET
print ('Normalized distance between ', audio_clip_s, ' and ', tar_name, ' is ', s_t_median) # 0 for similar audios
ws.append([audio_clip_s, tar_name, s_t_median])

#DISTANCE BETWEEN 1 SOURCE AND 1 MIMICKED
distX, cost, acc_cost, path = dtw(mfcc_s.T, mfcc_m.T, dist=lambda x, y: norm(x - y, ord=1))
print ('Normalized distance between ', audio_clip_s, ' and ', audio_clip_m, ' is ', distX) # 0 for similar audios
ws.append([audio_clip_s, audio_clip_m, distX])

#DISTANCE BETWEEN 1 MIMICKED AND ALL TARGET
print ('Normalized distance between ', audio_clip_m, ' and ', tar_name, ' is ', t_m_median) # 0 for similar audios
ws.append([audio_clip_m, tar_name, t_m_median, s_t_median - t_m_median])

wb.save(wb_path)
print("\n--- %s seconds ---" % (time.time() - start_time))
